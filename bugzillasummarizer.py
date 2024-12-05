import os
import openai
import pandas as pd
import time
from tqdm import tqdm
import re
import openpyxl

# Initialize the OpenAI client
openai.api_key = os.getenv("OPENAI_API_KEY")  # Ensure you have set the environment variable

# Function to parse the scraped data
def parse_scraped_data(file_path):
    bugs_data = []
    urls = []
    with open(file_path, 'r', encoding='utf-8') as file:
        current_bug = ""
        current_url = ""
        for line in file:
            if line.strip().startswith("LINK"):  # Start of new bug
                if current_bug:  # Save previous bug data
                    bugs_data.append(current_bug.strip())
                    urls.append(current_url.strip())
                current_bug = line
                current_url = line.split("LINK: ")[1].strip()
            else:
                current_bug += line
        if current_bug:  # Add the last bug data
            bugs_data.append(current_bug.strip())
            urls.append(current_url.strip())
    return bugs_data, urls

# Function to extract details from the response
def extract_bug_details(response_text, url):
    details = {"URL": url}
    sections = ["Bug Description", "Reproduction Steps", "Bug Type", "System Call Name", "Process/Application Interleaving", "Processes/Signals/Interrupts"]
    
    for i, section in enumerate(sections):
        if i < len(sections) - 1:
            next_section = sections[i + 1]
            pattern = re.compile(rf"{section}:(.*?)(?=\n{next_section}:|\Z)", re.DOTALL)
        else:
            pattern = re.compile(rf"{section}:(.*?)(?=\Z)", re.DOTALL)
        
        match = re.search(pattern, response_text)
        details[section] = match.group(1).strip() if match else "N/A"
    return details

# Create a folder for the Excel files if it doesn't exist
output_folder = "excel_files"
os.makedirs(output_folder, exist_ok=True)

# Collect all responses for the final combined file
all_responses = []

# Collect responses and process the data files in batches
batch_number = 1
while True:
    try:
        # Load and parse the file
        file_path = f"scraped_data_batch_{batch_number}.txt"
        bugs_data, urls = parse_scraped_data(file_path)

        # Check if there are no more files to process
        if not bugs_data:
            print("No more files to process.")
            break

        responses = []

        # Split bugs_data and urls into sub-batches
        sub_batch_size = 5  # Adjust the size to avoid exceeding token limits
        for i in range(0, len(bugs_data), sub_batch_size):
            sub_bugs_data = bugs_data[i:i + sub_batch_size]
            sub_urls = urls[i:i + sub_batch_size]

            # Send each sub-batch to GPT and process response with a progress bar
            for bug_data, url in tqdm(zip(sub_bugs_data, sub_urls), desc=f"Processing Batch {batch_number}, Sub-batch {i // sub_batch_size + 1}", total=len(sub_bugs_data)):
                try:
                    response = openai.ChatCompletion.create(
                        model="gpt-4o-mini",  # Corrected model name as per your request
                        messages=[
                            {
                                "role": "system",
                                "content": "Give results in this way only:\nBug Description:\nReproduction Steps\nBug Type\nSystem Call Name\nProcess/Application Interleaving\nProcesses/Signals/Interrupts"
                            },
                            {
                                "role": "user",
                                "content": bug_data
                            }
                        ],
                        temperature=1,
                        max_tokens=2048,
                        top_p=1,
                        frequency_penalty=0,
                        presence_penalty=0
                    )

                    response_text = response['choices'][0]['message']['content']
                    bug_details = extract_bug_details(response_text, url)
                    responses.append(bug_details)
                    all_responses.append(bug_details)
                except openai.error.InvalidRequestError as e:
                    print(f"Error processing sub-batch {i // sub_batch_size + 1}: {e}")
                    continue
                except openai.error.APIError as e:
                    print(f"API error processing sub-batch {i // sub_batch_size + 1}: {e}")
                    continue
                except openai.error.RateLimitError as e:
                    print(f"Rate limit exceeded, retrying sub-batch {i // sub_batch_size + 1}: {e}")
                    time.sleep(10)  # Wait before retrying

        # Convert responses to DataFrame with separate columns
        df = pd.DataFrame(responses)

        # Replace empty cells with "N/A"
        df.fillna("N/A", inplace=True)

        # Save to Excel with retry mechanism and adjust column widths
        while True:
            try:
                excel_file_path = os.path.join(output_folder, f"extracted_bug_details_batch_{batch_number}.xlsx")
                with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Bug Details')

                    # Access the workbook and worksheet objects
                    workbook = writer.book
                    worksheet = writer.sheets['Bug Details']

                    # Set the column widths and wrap text to avoid overlap
                    for i, col in enumerate(df.columns):
                        max_length = max(df[col].astype(str).map(len).max(), len(col))  # Get max length of column content
                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = max_length + 2  # Add some padding
                        for cell in worksheet[openpyxl.utils.get_column_letter(i + 1)]:
                            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Enable text wrapping

                print(f"Extraction and summarization complete. Saved in '{excel_file_path}'")
                break
            except PermissionError:
                input(f"The file '{excel_file_path}' is open. Please close it and press Enter to try again.")

        # Sleep for 6 seconds before processing the next batch
        time.sleep(6)

        # Increment the batch number
        batch_number += 1

    except FileNotFoundError:
        print(f"No more files found for batch {batch_number}.")
        break

# Save the combined data to a final Excel file
df_all = pd.DataFrame(all_responses)
df_all.fillna("N/A", inplace=True)
combined_excel_file_path = os.path.join(output_folder, "combined_bug_details.xlsx")
with pd.ExcelWriter(combined_excel_file_path, engine='openpyxl') as writer:
    df_all.to_excel(writer, index=False, sheet_name='Bug Details')

    # Access the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Bug Details']

    # Set the column widths and wrap text to avoid overlap
    for i, col in enumerate(df_all.columns):
        max_length = max(df_all[col].astype(str).map(len).max(), len(col))  # Get max length of column content
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = max_length + 2  # Add some padding
        for cell in worksheet[openpyxl.utils.get_column_letter(i + 1)]:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Enable text wrapping

print(f"Final combined extraction and summarization complete. Saved in '{combined_excel_file_path}'")
